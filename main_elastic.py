#!/usr/bin/python
from openpyxl import Workbook, load_workbook
from time_elastic import timelist
from class_elastic_request import elastic_request
from styles_elasticv2 import styles_line_hor, styles_line_vert, a_merge, draw, generate_cell_addr
from datetime import datetime
from crutch_elastic import fck_decorator
import args_parse_elastic
from elastic_enrich import elastic_request_enrich
from host_class import host_class
from notnorm_body import matrix_notnorm


@fck_decorator
class matrix:
    def decorate(self):
        pass

    def __init__(self):
        self.t = timelist()
        matrix_notnorm.t=self.t
        #print(123, matrix_notnorm.t)
        self.spisok = []
        self.time_len = len(self.t)
        print(f"There are {self.t.__len__() - 1} one-hour intervals\n")
        print("Successfull start script")


        for j in range(self.time_len):
            self.time_from = self.t.lst[j]
            if j + 1 < self.time_len:
                self.time_to = self.t.lst[j + 1]
                self.spisok.append(elastic_request("http://"+args_parse_elastic.arguments.url,self.time_from, self.time_to))

        self.hostname_vendor_dict = dict()
        self.hostname_vendor_list_default = []
        #hostname_vendor_ok = []
        #print(self.time_from,self.time_to)
        a = elastic_request("http://"+args_parse_elastic.arguments.url,self.t.lst[0], self.t.lst[self.time_len - 1])
        self.js = a()
        del a
        self.len_rows = len(self.js)
        #print(self.len_rows)

        for i in range(self.len_rows):
            vendor = self.js[i][0]
            hostname = self.js[i][1]
            hv = f"{hostname}///{vendor}"
            self.hostname_vendor_list_default.append(hv)  # составление полного перечня хостов за весь период
            self.hostname_vendor_dict[hv] = [hv]

        self.set_host_vendor_default = set(self.hostname_vendor_list_default)

        for j in range(self.time_len - 1):  # пробегаем по списку с объектами mp_request, вызываем каждый для запроса
            js = self.spisok[j]()
            hostname_vendor_ok = []
            for i in range(len(js)):  # от 0 до кол-ва хостов за все время
                count_events = round(float((js[i][2]) / 3600),4)  # забираем количество событий с i объекта
                hostname = js[i][1]
                vendor = js[i][0]
                hv = f"{hostname}///{vendor}"
                if hv in self.hostname_vendor_dict:
                    self.hostname_vendor_dict[hv].append(count_events)  # 123
                    hostname_vendor_ok.append(hv)
                #print(hostname_vendor_ok)


            set2_ok = set(hostname_vendor_ok)
            list2_ne_ok = list(self.set_host_vendor_default - set2_ok)

            for k in list2_ne_ok:
                self.hostname_vendor_dict[k].append(0)
        del self.spisok
        self.hosts2 = []
        for i, v in enumerate(self.hostname_vendor_dict.items()):
            self.hosts2.append(host_class(v[0].split("///")[0], v[1:]))
            ven = v[0].split("///")[1]
            self.hosts2[i].vendor = ven
            #print("lst1", v[1:])

        del self.hostname_vendor_dict


        enrich_request=elastic_request_enrich("http://"+args_parse_elastic.arguments.url,self.t.lst[0], self.t.lst[self.time_len - 1])()
        #print(enrich_request)
        for i in enrich_request:
            for j in self.hosts2:
                if j.vendor==i[0] and j.hostname==i[1]:
                    j.tag=i[2]
                    j.title=i[3]
                    #print(j.tag,j.title)


        '''тут можно вставлять костыли =)'''
        self.decorate()  # Выравнивает кривые данные сырого запроса и агрегирует в итоговой таблице
        """конец костылей"""

        self.hosts2 = sorted(self.hosts2, key=lambda x: x.hostname)
        self.hosts2 = sorted(self.hosts2, key=lambda x: x.title)
        self.hosts2 = sorted(self.hosts2, key=lambda x: x.vendor)
        self.hosts2 = self.hosts2 + matrix_notnorm()()

        self.rez_list = []
        for i in range(0, self.time_len - 1):
            summ = 0
            for j in self.hosts2:
                #print(len(j.lst), j.lst, j.hostname)
                #print(j.vendor,j.hostname,j.lst)
                if j.lst[i] != None and type(j.lst[i]) in (int, float):
                    summ += j.lst[i]
                #  print(j.lst[i])
            self.rez_list.append(summ)

        self.rez_list2 = []
        self.len_hosts2 = len(self.hosts2)
        for i in range(0, self.time_len-1):
            gen = generate_cell_addr(i+5)
            self.rez_list2.append(f"=SUM({gen}{2}:{gen}{self.len_hosts2+1})")

        self.title_tmp = self.t.title
        self.title_tmp.pop(0)
        self.str_title = ["event_src.vendor", "event_src.title", "event_src.host"] + ["tag"] + self.title_tmp  + ["Max value"] + ["Avarage value"]
        self.max_rez2 = f"=MAX(E{self.len_hosts2 + 2}:{generate_cell_addr(self.time_len + 3)}{self.len_hosts2 + 2})"
        self.avg_rez2 = f"=ROUND(AVERAGE(E{self.len_hosts2 + 2}:{generate_cell_addr(self.time_len + 3)}{self.len_hosts2 + 2}),5)"
        self.str_last2 = ["Total", f"{self.len_hosts2} assets","",""] + self.rez_list2 + [self.max_rez2, self.avg_rez2]
        #print("last", self.str_last2)
        #self.str_last2.pop(4)

        self.width = self.time_len + 6  # логично, так как к списку добавлено 4 поля спереди и 2 поля сзади
        self.length = len(self.hosts2) + 2
        #print(self.width, self.length)
        #print("matrix", self.__dict__)


def def_excel():
    Matrix = matrix()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("События")
    ws.append(Matrix.str_title)

    for i, v in enumerate(Matrix.hosts2):
        v.number = i + 2
        #print("aaa",Matrix.hosts2[i]())
        ws.append(Matrix.hosts2[i]())

    ws.append(Matrix.str_last2)

    #print(Matrix.width, Matrix.length)
    Matrix.length=len(Matrix.hosts2)+2
    del Matrix.hosts2
    draw(wb, ws, Matrix.width, Matrix.length, Matrix.rez_list)

    print("\nCreating xls report file in script directory")

    filename = (f'{args_parse_elastic.arguments.url}'
                f'_{datetime.now().strftime('%Y-%m-%d-%H-%M')}.xlsx')
    try:
        wb.save(filename)
    except PermissionError:
        print("\n!!!")
        print("To save the results, close the Excel file")

    print("Successfull finish")

def_excel()

#Matrix = matrix()

#for i in Matrix():
#    print(i())