#!/usr/bin/python
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )
from openpyxl.chart import LineChart, Reference
import openpyxl


def a_merge(ws,height):
    segment = []
    x = ws[f"A{2}"].value
    for k in range(2,height):
        tmp=ws[f"A{k}"].value
        if x==tmp:
            segment.append(k)
        else:
            #print(segment)
            ws.merge_cells(f"A{segment[0]}:A{segment[-1]}")
            segment=[]
            x=ws[f"A{k}"].value
            segment.append(k)
        #print(x)
    #print(segment)
    ws.merge_cells(f"A{segment[0]}:A{segment[-1]}")

def style(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal='left', vertical='top')

def generate_cell_addr(Numb):
    if Numb<1:
        return ''
    alfavit = ['Z','A','B',"C",'D','E','F','G','H','I',"J",'K',"L",'M','N','O','P','Q','R','S','T','U','V','W','X','Y']
    ost = alfavit[Numb%26]
    if (Numb//26)>0: #and not(((Numb//26)==1) and (Numb%26 == 0))
        if ost=='Z':
            ost = generate_cell_addr((Numb//26)-1)+ost
        else:
            ost = generate_cell_addr((Numb//26))+ost
    return ost


def styles_line_hor(ws,length,str_num):    #length-ширира (по времени)    #str-num - высота
    thin = Side(border_style="thick")
    for i in range(1,length):
        num=f'{generate_cell_addr(i)}{str_num}'
        #print(num)
        top_string = ws[num]
        top_string.fill = PatternFill('solid', fgColor="7CFC00")
        top_string.border = Border(bottom=thin,top=thin,right=thin,left=thin)
        top_string.font=Font(bold=True)



def styles_line_vert(ws,height,width=1):
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 31
    ws.column_dimensions['D'].width = 15
    cell = generate_cell_addr(width)
    #print(111,cell)
    ws.column_dimensions[cell].width = 17
    cell = generate_cell_addr(width+1)
   # print(222,cell)
    ws.column_dimensions[cell].width = 17
    thin = Side(border_style="thick")
    for i in range(1, height):
        top_string = ws[f'A{i}']
        top_string.border = Border(bottom=thin,top=thin,right=thin,left=thin)
        top_string.fill = PatternFill('solid', fgColor="87CEFA")
        top_string = ws[f'B{i}']
        top_string.border = Border(bottom=thin,top=thin,right=thin,left=thin)
        top_string.fill = PatternFill('solid', fgColor="87CEFA")
        top_string = ws[f'C{i}']
        top_string.border = Border(bottom=thin, top=thin, right=thin, left=thin)
        top_string.fill = PatternFill('solid', fgColor="87CEFA")
        top_string = ws[f'D{i}']
        top_string.border = Border(bottom=thin, top=thin, right=thin, left=thin)
        top_string.fill = PatternFill('solid', fgColor="87CEFA")


        cell=generate_cell_addr(width)
        top_string = ws[f'{cell}{i}']
        top_string.border = Border(bottom=thin,top=thin,right=thin,left=thin)
        top_string.fill = PatternFill('solid', fgColor="FFFF00")
        cell = generate_cell_addr(width+1)
        top_string = ws[f'{cell}{i}']
        top_string.border = Border(bottom=thin,top=thin,right=thin,left=thin)
        top_string.fill = PatternFill('solid', fgColor="FFFF00")


def diagram(wb,width,height,rez_list):
    start=generate_cell_addr(5)
    stop=generate_cell_addr(width-3)
    vertikal=height
    #print(width,height)
    sheet=wb["События"]
    ws = wb.create_sheet("Диаграмма")
    str1=sheet[f"{start}{1}:{stop}{1}"]
    #str2=sheet[f"{start}{vertikal}:{stop}{vertikal}"]
    #print(f"Диаграмма считается по {start}{vertikal}:{stop}{vertikal}, список {str2[0]}")
    lst1=[x.value for x in str1[0]]
    #lst2=[x.value for x in str2[0]]
    lst2=rez_list
    # print("sdf",sheet[f"{start}{vertikal}"].value)
   # print("this list",lst2)
    z_tuple=zip(lst1,lst2)
    z_list=list(z_tuple)
   # print(z_list)
    for i in range(len(lst1)):
        ws.append(z_list[i])
    values = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(lst1))
    #ws['B1']=ws['B1'].value
    date_cat = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=len(lst1))
    chart = LineChart()
    chart.add_data(values)
    chart.set_categories(date_cat)
    chart.title = "Показатели EPS"
    chart.style=10
    #chart.series[1].graphicalProperties.line.solidFill = "00AAAA"
    #chart.series[0].graphicalProperties.line.dashStyle = "sysDot"
    chart.series[0].graphicalProperties.line.width = "12050"
    chart.series[0].marker.symbol="auto"
    chart.series[0].marker.size=3
    chart.x_axis.title = "Дата"
    chart.y_axis.title = "EPS"
    chart.height = 11  # default is 7.5
    chart.width = 25
    ws.add_chart(chart, "E2")


def draw(wb,ws,width,height,rez_list):
    styles_line_vert(ws, height, width - 2)
    styles_line_hor(ws, width, 1)
    styles_line_hor(ws, width, height)  #
    a_merge(ws, height)
    style(ws)
    diagram(wb,width,height,rez_list)
